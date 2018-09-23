try:
    import openpyxl
except ImportError:
    print("you need to install openpyxl")
try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string


def open_excel_and_replace(path, old_str, new_str):
    wb = openpyxl.load_workbook(path)
    print(">>>>>>>>")
    print("wb",wb)
    print("<<<<<<<<<")
    for sheet in wb:
        max_col = get_column_letter(sheet.max_column)
        max_row = str(sheet.max_row)
        lastcell = max_col + max_row
        print("last cell: ", max_col + max_row)
        params = [old_str, new_str]
        searchFromAToBAndDo("A1", lastcell, replaceStringInCell, params, sheet)
        # print(sheet['A1'].value)

        wb.save(path)


def replaceStringInCell(cell, params):
    try:
        if (params != None and (isinstance(params, (list, tuple)))):

            if (len(params) != 2):
                raise Exception("lenght of params is not 2. It must be [oldstr , newstr]")

            old_str = params[0]
            new_str = params[1]
            tempValue = cell.value
            if (isinstance(tempValue, str)):
                cell.value = tempValue.replace(old_str, new_str)
        else:
            raise TypeError
    except TypeError:
        print("params is None or non-list type")


def searchFromAToBAndDo(A, B, callback, paramForCallBack, sheet):
    for row in sheet[A:B]:
        for cell in row:
            callback(cell, paramForCallBack)
